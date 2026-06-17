"""
One-time build tool: parse the PNE Pizza training workbook into a clean,
structured JSON fixture (database/data/training_content.json) that
TrainingContentSeeder consumes. Run from the project root:

    python3 database/data/build_training_content.py

The workbook is positionally encoded, so parsing is driven by per-sheet
configuration below. Re-run whenever the workbook is revised.
"""

import json
import re
from pathlib import Path

import openpyxl

ROOT = Path(__file__).resolve().parents[2]
WORKBOOK = ROOT / "Plan Phase 1 (PNE  FOODS's conflicted copy 2026-06-15).xlsx"
OUTPUT = ROOT / "database" / "data" / "training_content.json"

VERB = re.compile(
    r"^(Demonstrate|Explain|Verify|Show|Describe|Identify|List|State)\b", re.I
)

# Section presentation: display order + lucide icon name (matches frontend).
SECTION_META = {
    "Introduction": (0, "BookOpen"),
    "Product Prep": (1, "Boxes"),
    "Pressouts": (2, "Croissant"),
    "Dough & Sauce": (3, "ChefHat"),
    "Pizza Dress (Making)": (4, "Pizza"),
    "Landing": (5, "UtensilsCrossed"),
    "Dishwashing": (6, "Droplets"),
    "Customer Service": (7, "ShoppingCart"),
}

# Checklist sheets: which column holds item text + the known category headers.
SHEETS = {
    "ProductPrep - Equipment": {
        "section": "Product Prep",
        "text_col": "C",
        "headers": ["Topping Prep", "Sauce Prep"],
    },
    "Pressout": {
        "section": "Pressouts",
        "text_col": "D",
        "headers": ["Round & Crazy Bread Pressouts"],
    },
    "Dough Prep&Sauce": {
        "section": "Dough & Sauce",
        "text_col": "C",
        "headers": ["Round & Crazy Bread dough", "Deep Dish Dough", "Sauce"],
    },
    "Making": {
        "section": "Pizza Dress (Making)",
        "text_col": "C",
        "headers": [
            "Round Sauce, Cheese & Pepperoni",
            "Deep Dish",
            "Custom & Specialty Pizzas",
            "Crazy Bread",
        ],
    },
    "Landing": {
        "section": "Landing",
        "text_col": "D",
        "headers": ["Pizza Landing", "Crazy bread & Sides Landing", "Quiz"],
    },
    "Dishwashing": {
        "section": "Dishwashing",
        "text_col": "C",
        "headers": [],
    },
    "CS - Register": {
        "section": "Customer Service",
        "text_col": "D",
        "headers": ["Front Counter", "Telephone & Drive-Thru (Pick-Up Window)"],
    },
}

# Timeline B-column label -> our section title.
TIMELINE_NAME_TO_SECTION = {
    "Introduction": "Introduction",
    "Pressouts": "Pressouts",
    "Product Prep": "Product Prep",
    "Pizza Dress": "Pizza Dress (Making)",
    "Dough": "Dough & Sauce",
    "Dishwashing": "Dishwashing",
    "Landing": "Landing",
    "Customer Service": "Customer Service",
}

SKIP_TEXT = {"Trainer's Checklist", "Trainers Checklist", ""}
DEFAULT_CATEGORY = "Station Overview"


def cell_text(value):
    return str(value).strip() if value is not None else ""


def collect_links(values):
    links = []
    for v in values.values():
        if isinstance(v, str):
            s = v.strip()
            if s.lower().startswith("http"):
                links.append(s)
    return links


def parse_timeline(wb):
    ws = wb["Timeline"]
    timing = {}
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        vals = {c.column_letter: c.value for c in row}
        name = cell_text(vals.get("B"))
        section = TIMELINE_NAME_TO_SECTION.get(name)
        if not section:
            continue
        timing[section] = {
            "pie_content_review": cell_text(vals.get("C")) or None,
            "screen_to_shoulder": cell_text(vals.get("D")) or None,
            "hands_on_shifts": cell_text(vals.get("E")) or None,
        }
    return timing


def new_section(title):
    order, icon = SECTION_META[title]
    return {
        "title": title,
        "icon": icon,
        "order": order,
        "pie_content_review": None,
        "screen_to_shoulder": None,
        "hands_on_shifts": None,
        "categories": [],
    }


def parse_checklist_sheet(ws, cfg):
    section = new_section(cfg["section"])
    section_title = cfg["section"]
    headers = set(cfg["headers"])
    text_col = cfg["text_col"]

    categories = {}  # title -> dict
    order_counter = {"cat": 0, "item": {}}
    current_cat = None
    current_item = None

    def ensure_category(name):
        nonlocal current_cat
        if name not in categories:
            cat = {"title": name, "order": order_counter["cat"], "items": []}
            order_counter["cat"] += 1
            order_counter["item"][name] = 0
            categories[name] = cat
            section["categories"].append(cat)
        current_cat = categories[name]
        return current_cat

    def attach_links(item, links):
        for url in links:
            item["media"].append({"type": "link", "url": url, "label": None})

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row):
        values = {c.column_letter: c.value for c in row if c.value is not None}
        if not values:
            continue
        text = cell_text(values.get(text_col))
        links = collect_links(values)
        has_bool = any(isinstance(v, bool) for v in values.values())

        # Meta rows (section title, "Trainer's Checklist", blanks).
        if text in SKIP_TEXT or text == section_title:
            if current_item is not None:
                attach_links(current_item, links)
            continue

        if text in headers:
            ensure_category(text)
            current_item = None
            continue

        is_item = has_bool or bool(VERB.match(text))
        if is_item:
            if current_cat is None:
                ensure_category(DEFAULT_CATEGORY)
            name = current_cat["title"]
            current_item = {
                "title": text,
                "content": None,
                "_content_lines": [],
                "importance": "highly_important",
                "order": order_counter["item"][name],
                "media": [],
                "children": [],
            }
            order_counter["item"][name] += 1
            current_cat["items"].append(current_item)
            attach_links(current_item, links)
        else:
            # Expected answer / explanatory content for the preceding item.
            if current_item is not None:
                current_item["_content_lines"].append(text)
                attach_links(current_item, links)

    # Finalize content.
    for cat in section["categories"]:
        for item in cat["items"]:
            lines = item.pop("_content_lines")
            item["content"] = "\n".join(lines) if lines else None
    return section


def parse_introduction(wb):
    ws = wb["Introduction"]
    section = new_section("Introduction")
    cat = {"title": "Orientation", "order": 0, "items": []}
    section["categories"].append(cat)

    rows = {c.coordinate: c.value for row in ws.iter_rows() for c in row if c.value is not None}
    specs = [
        ("Get to know them", None, "B3"),
        ("Every Day Routine", "C4", "B4"),
        ("Basic Store Cleaning Instructions and Food Safety", "C5", "B5"),
        ("Pizza Life Cycle in General", "C6", "B6"),
    ]
    order = 0
    for title, content_cell, _label in specs:
        content = cell_text(rows.get(content_cell)) if content_cell else ""
        # Drop local file-path "content" that isn't useful to trainees.
        if content.lower().startswith("c:") or content.lower().startswith("drop"):
            content = ""
        cat["items"].append(
            {
                "title": title,
                "content": content or None,
                "importance": "highly_important",
                "order": order,
                "media": [],
                "children": [],
            }
        )
        order += 1
    return section


def main():
    wb = openpyxl.load_workbook(WORKBOOK, data_only=True)
    timing = parse_timeline(wb)

    sections = [parse_introduction(wb)]
    for sheet, cfg in SHEETS.items():
        sections.append(parse_checklist_sheet(wb[sheet], cfg))

    # Apply timing metadata and drop empty categories (workbook headers
    # occasionally have no items beneath them).
    for section in sections:
        if section["title"] in timing:
            section.update(timing[section["title"]])
        section["categories"] = [c for c in section["categories"] if c["items"]]
        for order, cat in enumerate(section["categories"]):
            cat["order"] = order

    sections.sort(key=lambda s: s["order"])

    OUTPUT.write_text(json.dumps({"sections": sections}, indent=2, ensure_ascii=False))

    # Summary for sanity-checking.
    total_items = 0
    for s in sections:
        cats = s["categories"]
        items = sum(len(c["items"]) for c in cats)
        total_items += items
        print(f"{s['title']:<24} cats={len(cats):<2} items={items}")
    print(f"{'TOTAL':<24} sections={len(sections)} items={total_items}")
    print(f"Wrote {OUTPUT.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
